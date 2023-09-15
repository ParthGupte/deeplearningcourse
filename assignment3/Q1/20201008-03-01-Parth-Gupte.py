import os
import numpy as np
import openpyxl as pyxl

def read_conditions(file_name = "asst-3-Q1.txt"):
    with open(os.path.dirname(__file__)+"/"+file_name,'r') as f:
        lines = f.readlines()
        cond_lst = []
        m , n = [int(x) for x in lines[0].split()]
        for line in lines[1:]:
            str_lst = line.split()
            cond = [str_lst[0].lower(),float(str_lst[1]),float(str_lst[2])]
            cond_lst.append(cond)
    
    return cond_lst

def read_data(file_name = "asst-3-Q1.xlsx"):
    wb_obj = pyxl.load_workbook(os.path.dirname(__file__)+"/"+file_name)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    data_arr = np.empty((m_row,m_col))
    for i in range(m_row):
        for j in range(m_col):
            cell_obj = sheet_obj.cell(row = i+1, column = j+1)
            data_arr[i,j] = cell_obj.value
    return data_arr

def fix_row(row,cond):
    #checking range
    l, u = cond[1:]
    for x in row:
        if x == np.nan:
            break
        elif x<l or x>u:
            break
        elif cond[0] == 'int':
            if int(x) != x:
                break
        elif cond[0] == 'float':
            if float(x) != x:
                break
    else:
        if cond[0] == 'int':
            return [int(x) for x in row]
        elif cond[0] == 'float':
            return [float(x) for x in row]
        
    return []
    
def fix_data(data:np.ndarray):
    '''
    Input an array and a cleaned array is returned
    '''
    new_data_lst = []
    for i in range(len(data)):
        row = list(data[i])
        new_row = fix_row(row,cond_lst[i])
        if new_row != []:
            new_data_lst.append(new_row)
    return np.array(new_data_lst)


cond_lst = read_conditions()
data = read_data()
new_data = fix_data(data)
print("20201008", len(new_data))