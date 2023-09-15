"""
network.py
~~~~~~~~~~

A module to implement the stochastic gradient descent learning
algorithm for a feedforward neural network.  Gradients are calculated
using backpropagation.  Note that I have focused on making the code
simple, easily readable, and easily modifiable.  It is not optimized,
and omits many desirable features.
"""

#### Libraries
# Standard library
import random
import openpyxl as pyxl
import os

# Third-party libraries
import numpy as np

class Network(object):

    def __init__(self, sizes):
        """The list ``sizes`` contains the number of neurons in the
        respective layers of the network.  For example, if the list
        was [2, 3, 1] then it would be a three-layer network, with the
        first layer containing 2 neurons, the second layer 3 neurons,
        and the third layer 1 neuron.  The biases and weights for the
        network are initialized randomly, using a Gaussian
        distribution with mean 0, and variance 1.  Note that the first
        layer is assumed to be an input layer, and by convention we
        won't set any biases for those neurons, since biases are only
        ever used in computing the outputs from later layers."""
        self.num_layers = len(sizes)
        self.sizes = sizes
        self.biases = [np.random.randn(y, 1) for y in sizes[1:]]
        self.weights = [np.random.randn(y, x)
                        for x, y in list(zip(sizes[:-1], sizes[1:]))]


    def feedforward(self, a):
        """Return the output of the network if ``a`` is input."""
        for b, w in list(zip(self.biases, self.weights)):
            a = sigmoid(np.dot(w, a)+b)
        return a

    def SGD(self, training_data, epochs, mini_batch_size, eta,
            test_data=None):
        """Train the neural network using mini-batch stochastic
        gradient descent.  The ``training_data`` is a list of tuples
        ``(x, y)`` representing the training inputs and the desired
        outputs.  The other non-optional parameters are
        self-explanatory.  If ``test_data`` is provided then the
        network will be evaluated against the test data after each
        epoch, and partial progress printed out.  This is useful for
        tracking progress, but slows things down substantially."""
        #test_data = list(zip(test_data))
        if test_data: n_test = len(test_data)
        #training_data = list(zip(training_data))
        n = len(training_data)
        for j in range(epochs):
            random.shuffle(training_data)
            mini_batches = [
                training_data[k:k+mini_batch_size]
                for k in range(0, n, mini_batch_size)]
            for mini_batch in mini_batches:
                self.update_mini_batch(mini_batch, eta, test_data)
            if test_data:
                print("Epoch {0}: {1} / {2}".format(
                    j, self.evaluate(test_data), n_test))
            else:
                print("Epoch {0} complete".format(j))

    def update_mini_batch(self, mini_batch, eta, test_data):
        """Update the network's weights and biases by applying
        gradient descent using backpropagation to a single mini batch.
        The ``mini_batch`` is a list of tuples ``(x, y)``, and ``eta``
        is the learning rate."""
        nabla_b = [np.zeros(b.shape) for b in self.biases]
        nabla_w = [np.zeros(w.shape) for w in self.weights]
        for x, y in mini_batch:
            delta_nabla_b, delta_nabla_w = self.backprop(x, y)
            nabla_b = [nb+dnb for nb, dnb in list(zip(nabla_b, delta_nabla_b))]
            nabla_w = [nw+dnw for nw, dnw in list(zip(nabla_w, delta_nabla_w))]
        self.weights = [w-(eta/len(mini_batch))*nw
                        for w, nw in list(zip(self.weights, nabla_w))]
        self.biases = [b-(eta/len(mini_batch))*nb
                       for b, nb in list(zip(self.biases, nabla_b))]
        #print("During Epoch", self.evaluate(test_data), len(test_data))

    def backprop(self, x, y):
        """Return a tuple ``(nabla_b, nabla_w)`` representing the
        gradient for the cost function C_x.  ``nabla_b`` and
        ``nabla_w`` are layer-by-layer lists of numpy arrays, similar
        to ``self.biases`` and ``self.weights``."""
        nabla_b = [np.zeros(b.shape) for b in self.biases]
        nabla_w = [np.zeros(w.shape) for w in self.weights]
        # feedforward
        activation = x
        activations = [x] # list to store all the activations, layer by layer
        zs = [] # list to store all the z vectors, layer by layer
        for b, w in list(zip(self.biases, self.weights)):
            z = np.dot(w, activation)+b
            zs.append(z)
            activation = sigmoid(z)
            activations.append(activation)
        # backward pass
        delta = self.cost_derivative(activations[-1], y) * \
            sigmoid_prime(zs[-1])
        nabla_b[-1] = delta
        nabla_w[-1] = np.dot(delta, activations[-2].transpose())
        # Note that the variable l in the loop below is used a little
        # differently to the notation in Chapter 2 of the book.  Here,
        # l = 1 means the last layer of neurons, l = 2 is the
        # second-last layer, and so on.  It's a renumbering of the
        # scheme in the book, used here to take advantage of the fact
        # that Python can use negative indices in lists.
        for l in range(2, self.num_layers):
            z = zs[-l]
            sp = sigmoid_prime(z)
            delta = np.dot(self.weights[-l+1].transpose(), delta) * sp
            nabla_b[-l] = delta
            nabla_w[-l] = np.dot(delta, activations[-l-1].transpose())
        return (nabla_b, nabla_w)

    def evaluate(self, test_data):
        """Return the number of test inputs for which the neural
        network outputs the correct result. Note that the neural
        network's output is assumed to be the index of whichever
        neuron in the final layer has the highest activation."""
        test_results = [(np.argmax(self.feedforward(x)), y)
                        for (x, y) in test_data]
        return sum(int(x == y) for (x, y) in test_results)

    def cost_derivative(self, output_activations, y):
        """Return the vector of partial derivatives \partial C_x /
        \partial a for the output activations."""
        return (output_activations-y)

#### Miscellaneous functions
def sigmoid(z):
    """The sigmoid function."""
    return 1.0/(1.0+np.exp(-z))

def sigmoid_prime(z):
    """Derivative of the sigmoid function."""
    return sigmoid(z)*(1-sigmoid(z))



def read_data(file_name = "asst-3-Q3.xlsx"):
    wb_obj = pyxl.load_workbook(os.path.dirname(__file__)+"/"+file_name)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column
    Y = np.zeros((m_row,1))
    data_arr = np.zeros((13,m_row))
    for i in range(m_row):
        if i == 0:
            continue
        for j in range(m_col):
            cell_obj = sheet_obj.cell(row = i+1, column = j+1)
            v = cell_obj.value
            if j == 0:
                continue
            elif j == 1:
                data_arr[0,i-1] = v
            elif j == 2:
                if v == "France":
                    data_arr[1,i-1] = 1
                elif v == "Spain":
                    data_arr[2,i-1] = 1
                elif v == "Germany":
                    data_arr[3,i-1] = 1
            elif j == 3:
                if v == "Male":
                    data_arr[4,i-1] = 1
                else:
                    data_arr[5,i-1] = 1
            elif j == 4:
                data_arr[6,i-1] = v
            elif j == 5:
                data_arr[7,i-1] = v
            elif j == 6:
                data_arr[8,i-1] = v
            elif j == 7:
                data_arr[9,i-1] = v
            elif j == 8:
                data_arr[10,i-1] = v
            elif j == 9:
                data_arr[11,i-1] = v
            elif j == 10:
                data_arr[12,i-1] = v
            elif j == 11:
                Y[i-1,0] = v
            else:
                continue 

    return data_arr,Y

# data, Y = read_data()
# print(np.array([data[:,0]]).transpose().shape,Y[0])

def normalise(X: np.ndarray,row_wise = True):
    if not row_wise:
        X = X.transpose()
    mean = np.mean(X,axis=1)
    mean_arr = np.array([mean]*X.shape[1]).transpose()
    std = np.std(X,axis=1)
    std_arr = np.array([std]*X.shape[1]).transpose()
    N = (X-mean_arr)/std_arr
    if not row_wise:
        N = N.transpose()
    return N, mean_arr, std_arr


def preprocessing(data:np.ndarray):
    data_copy = data.copy()
    X = data_copy[[0,6,7,8,9,12]]
    N, mean_arr, std_arr = normalise(X)
    data_copy[[0,6,7,8,9,12]] = N
    return data_copy

def get_training_tuples(processed_data:np.ndarray,Y:np.ndarray):
    training_data = []
    for i in range(processed_data.shape[1]):
        x = np.array([processed_data[:,0]]).transpose()
        y = Y[i]
        training_data.append((x,y))
    return training_data

# data,Y = read_data()
# new_data = preprocessing(data)
# training_data = get_training_tuples(new_data,Y)

# churn_predictor = Network([13,10,1])
# w = churn_predictor.weights.copy()[0] 
# churn_predictor.SGD(training_data[:7500],500,100,1,[(x,y[0]) for x,y in training_data[7500:]])
# w_new = churn_predictor.weights[0]
# print((w_new == w).all())

# print(churn_predictor.sizes,churn_predictor.weights,churn_predictor.biases)

# nn1 = Network([2,3,1])
# # print(nn1.weights)
# x = np.array([[1],[2]])
# o = nn1.feedforward(x)
# print(o)
# nn1.backprop(x,np.array([[1]]))
# o = nn1.feedforward(x)
# print(o)



sizes, weights, biases = [13, 10, 1], [np.array([[ 0.27254611,  0.40140164,  0.46172411,  0.05837055, -0.18236724,
        -0.4542134 , -1.23959383, -1.16507112, -0.85364742, -1.12514415,
        -1.92912563, -1.54315202, -1.3907501 ],
       [-0.07452997,  0.13493272,  0.82066837,  0.55724783, -2.04224315,
        -0.62040875,  0.29307028, -0.38823592,  2.0460461 , -0.49965568,
        -0.26964588,  0.04296393,  0.62205011],
       [ 0.4752995 , -0.59564744, -0.17943499,  0.26323506,  0.27328795,
        -0.40690091, -0.10626259, -0.56082776, -0.4679703 , -0.24361397,
        -1.19105717,  0.51410498,  1.54703936],
       [ 0.06652008,  0.40805383, -0.04436841,  0.89756216, -1.736829  ,
         1.60228894, -0.40959762,  0.09876553,  0.02781981,  0.66811099,
        -0.30602923, -3.19987579,  1.04141503],
       [-0.10702788,  1.30592651, -0.76518511, -1.71719285, -0.49007639,
         0.48945513,  1.18838788,  1.54472494, -0.76966427,  0.60977213,
        -0.32978862,  0.78687683, -0.38682451],
       [ 0.30577513, -1.91687357,  0.18005311, -0.91212671, -0.60503473,
        -0.3219523 ,  0.21336757,  0.61321848, -2.17606005,  0.1977774 ,
        -2.42630682,  0.41841458,  1.2669648 ],
       [-1.67060271, -0.08736329, -0.4608976 ,  1.03971354,  0.5255432 ,
        -0.58499719, -1.24512044,  0.21134737, -1.26724978, -1.60442107,
         1.28163875, -0.80086859, -0.71942986],
       [-1.00252727, -0.08511659, -0.68514176,  1.03186005, -0.98257176,
        -1.36650242,  1.12702389,  0.69419986,  0.7040151 ,  1.32332818,
         0.51423557, -0.30592678,  0.20235729],
       [-0.19664145, -1.17118383, -0.98080174, -0.5122704 ,  1.37873791,
        -0.10168181,  1.28808187, -0.44592235,  0.39667408, -0.12140951,
        -0.61519156,  0.3433564 , -0.39915816],
       [-0.14161828, -0.57112792,  0.06312368,  0.16822631,  0.55001509,
         0.18032485,  0.74032197,  1.10469522,  0.20161871,  0.78861494,
         0.70383981, -1.30062528, -1.22516693]]), np.array([[-0.33654705, -0.3497868 ,  0.11515942, -0.45799039, -0.04922063,
        -0.11527634, -0.79174009, -0.18893125,  0.30106266,  3.39262963]])], [np.array([[ 0.20528916],
       [ 2.17904058],
       [-0.30691179],
       [-0.10788375],
       [-0.36220142],
       [ 0.00240374],
       [ 2.2472778 ],
       [-0.24556974],
       [-0.58789817],
       [-0.71995124]]), np.array([[-0.3390068]])]

churn_predictor = Network(sizes)
churn_predictor.weights = weights
churn_predictor.biases = biases

test_data, Y = read_data("asst-3-Q3-test.xlsx") 
new_test_data = preprocessing(test_data)
test_data_tuples = get_training_tuples(new_test_data,Y)[-1000:]
acc = churn_predictor.evaluate(test_data_tuples)
print(20201008,sizes[0],sizes[1],sizes[2],acc,len(test_data_tuples))