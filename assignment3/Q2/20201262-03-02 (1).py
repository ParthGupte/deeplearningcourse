import numpy as np
import openpyxl
import os

#inputting the data
wb = openpyxl.load_workbook(os.path.dirname(__file__)+'/asst-3-Q2.xlsx')
dataf = wb.active
n = []
r,c = dataf.max_row, dataf.max_column
for rowi in range(1,r+1):
    T = []
    for columni in range(1,c+1):
        data = dataf.cell(row = rowi , column = columni).value
        T.append(data)
    n.append(T)

n = np.array(n)
A = n

def Mean_vector():
    L = []
    for j in range(c):
        Sum = 0
        for i in range(r):
            Sum += n[i][j]
        L.append(Sum/r)
    L = list(np.mean(np.transpose(n),axis = 1))
    return L
X_mean = Mean_vector() 
    
def normallize1(X_mean):    
    X = []
    for i in range(r):
        X.append(X_mean)
    X = np.array(X)
    return X

X = normallize1(X_mean)
n = np.subtract(n,X)
def std_dev():
    L = np.std(np.transpose(n),axis = 1)
    return L

sigma_list = std_dev() 

def final_normallize(n):
    for i in range(r):
        for j in range(c):
            n[i][j] = (n[i][j])/sigma_list[j]

final_normallize(n)

m = np.matmul(np.transpose(n),n)
# finding eigenvalues and eigenvectors


w, v = np.linalg.eig(m)
S = []
r,c = v.shape

#Selecting Eigen_values with highest magnitude
E = [0,1]
if abs(w[E[0]]) < abs(w[E[1]]):
    E = [1,0]

for i in range(2,len(w)):
    e = w[i]
    if abs(e)> w[E[0]]:
        tmp = E[0]
        E[0] = i
        E[1] = tmp
    elif abs(e)> w[E[1]]:
        E[1] = i

#Creating the S matrix (eigen-value matrix)
for i in range(r):
    T = []
    for j in E:
        T.append(v[i][j])
    S.append(T)
S = np.array(S)
St = np.transpose(S)


B = np.matmul(n,S)
r,c = dataf.max_row,dataf.max_column
a = np.matmul(B,St)
def denormallize():
    for i in range(r):
        for j in range(c):
            a[i][j] = (a[i][j])*sigma_list[j]
    A2 = np.add(a,X)
    return A2

A2 = denormallize()
print(20201262,np.linalg.norm(np.subtract(A,A2))/(dataf.max_row*dataf.max_column),np.linalg.norm(B))









