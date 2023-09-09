import numpy as np
import openpyxl as pyxl
import os

def read_data(file_name = "asst-3-Q2.xlsx"):
    wb_obj = pyxl.load_workbook(os.path.dirname(__file__)+"/"+file_name)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    data_arr = np.empty((m_row,m_col))
    for i in range(m_row):
        for j in range(m_col):
            cell_obj = sheet_obj.cell(row = i+1, column = j+1)
            data_arr[i,j] = cell_obj.value
    return data_arr

def normalise(X: np.ndarray,row_wise = True):
    if not row_wise:
        X = X.transpose()
    mean = np.mean(X,axis=1)
    mean_arr = np.array([mean]*X.shape[1]).transpose()
    std = np.std(X,axis=1)
    std_arr = np.array([std]*X.shape[1]).transpose()
    N = (X-mean_arr)/std_arr
    if not row_wise:
        N = N.transpose()
    return N, mean_arr, std_arr

def denormalise(N: np.ndarray, mean_arr: np.ndarray, std_arr: np.ndarray):
    X = N*std_arr + mean_arr
    return X

def decrypt_martix(X:np.ndarray,l:int):
    N, mean_arr, std_arr = normalise(X)
    NTN = np.matmul(N.transpose(),N)
    eigvals, eigvecs = np.linalg.eig(NTN)
    idx_L = list(range(len(eigvals)))
    l_large_eig_idxs = sorted(idx_L,key = lambda x: abs(eigvals[x]), reverse=True)[:l]
    D = (eigvecs.transpose()[l_large_eig_idxs]).transpose()
    return D
    
def frobenius(M:np.ndarray):
    S = 0
    for i in range(M.shape[0]):
        for j in range(M.shape[1]):
            S += (M[i,j])**2
    S = (S**(1/2))#/(M.shape[0]*M.shape[1])
    return S

def encrypt(X:np.ndarray,D:np.ndarray):
    E = D.transpose()
    N, mean_arr, std_arr = normalise(X)
    return np.matmul(N,D), mean_arr, std_arr

def decrypt(B:np.ndarray,D:np.ndarray, mean_arr, std_arr):
    N = np.matmul(B,D.transpose())
    X = N*std_arr + mean_arr
    return X


X = read_data()
D = decrypt_martix(X,2)
B, mean_arr, std_arr = encrypt(X,D)
X_new = decrypt(B,D,mean_arr,std_arr)

print(20201008,frobenius(X-X_new)/(X.shape[0]*X.shape[1]),frobenius(B))

# normalise(X)